import sys
import subprocess
import logging
import os
import json
import math
import datetime
import tkinter as tk
from tkinter import ttk

# ============================================================
# Startup splash while libraries are loading
# ============================================================

def get_branding_icon():
    """Returns a reusable branded PNG icon for the splash and main window."""
    icon_data = """iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAAA/ElEQVR42s2Xyw3CMAyGPQEcuwMTsBzjMEanYAo2qBASlQ8BK3L8Kop9sKrUqf8vdvOC2+P9QTsvl+mGutAAZkM0TaCNCMTp/gyLfwGsEChmNYs4tkFyRoQlEC4+aIT4lAQsfmlwoKWJWh/Y6+fKC5Za/QuA03FloInQ9nXdRL87A94RUgBrhsIANIgEoH07BJD+chpoe/0MAWibK4c0PcEz3zkICjAq1yEAGqwXoQD9e22dUAG4GnIj7WcB14fLyiGAkYDmFwFG4jjCkUX69RB1MlDuH0ifBeXWgSkrYfpeUGI3TD8PlD4RTTkTpp6KU+8FqTej1Lth9u14Bxum5xw9DdIQAAAAAElFTkSuQmCC"""
    try:
        return tk.PhotoImage(data=icon_data)
    except Exception:
        return None


def show_startup_splash(message="Please wait, loading libraries..."):
    """Displays a branded splash window with a rotating network-style loading icon."""
    splash = tk.Tk()
    splash.withdraw()
    splash.title("mDNS Network Scanner")
    brand_icon = get_branding_icon()
    if brand_icon:
        try:
            splash.iconphoto(False, brand_icon)
        except Exception:
            pass
    splash.overrideredirect(True)
    splash.attributes("-topmost", True)
    splash.configure(bg="#0F172A")

    width = 440
    height = 190
    screen_width = splash.winfo_screenwidth()
    screen_height = splash.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    splash.geometry(f"{width}x{height}+{x}+{y}")

    style = ttk.Style(splash)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure("Splash.TFrame", background="#0F172A")
    style.configure("Splash.TLabel", background="#0F172A", foreground="#E2E8F0")

    container = ttk.Frame(splash, padding=(8, 10, 8, 8), style="Splash.TFrame")
    container.pack(fill=tk.BOTH, expand=True)

    header = tk.Frame(container, bg="#0F172A")
    header.pack(fill=tk.X)

    icon_label = tk.Label(header, bg="#0F172A", bd=0, padx=0, pady=0)
    icon_label.pack(side=tk.LEFT)
    if brand_icon:
        icon_label.configure(image=brand_icon)
        icon_label.image = brand_icon

    title = tk.Label(
        header,
        text="mDNS Network Scanner",
        font=("Segoe UI", 15, "bold"),
        bg="#0F172A",
        fg="#F8FAFC",
        anchor=tk.W,
    )
    title.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

    label = tk.Label(
        container,
        text=message,
        font=("Segoe UI", 10, "bold"),
        bg="#0F172A",
        fg="#E0F2FE",
        anchor=tk.CENTER,
        justify=tk.CENTER,
        wraplength=350,
    )
    label.pack(fill=tk.X, pady=(10, 0))

    canvas = tk.Canvas(container, width=110, height=96, bg="#0F172A", highlightthickness=0)
    canvas.pack(anchor=tk.CENTER, pady=(2, 0))

    angle = [0.0]

    def draw_network_icon():
        if not splash.winfo_exists():
            return

        canvas.delete("all")
        cx = 55
        cy = 55
        orbit_r = 28

        canvas.create_oval(cx - 36, cy - 36, cx + 36, cy + 36, outline="#1E293B", width=2)

        for i in range(3):
            theta = i * (2 * math.pi / 3) + angle[0]
            x = cx + orbit_r * math.cos(theta)
            y = cy + orbit_r * math.sin(theta)
            canvas.create_line(cx, cy, x, y, fill="#38BDF8", width=2)
            canvas.create_oval(x - 7, y - 7, x + 7, y + 7, fill="#7DD3FC", outline="#E0F2FE", width=1)

        canvas.create_oval(cx - 15, cy - 15, cx + 15, cy + 15, fill="#0EA5E9", outline="#E2E8F0", width=2)

        lead_a = angle[0]
        led_x = cx + orbit_r * math.cos(lead_a)
        led_y = cy + orbit_r * math.sin(lead_a)
        canvas.create_oval(led_x - 8, led_y - 8, led_x + 8, led_y + 8, fill="#F8FAFC", outline="#38BDF8", width=2)

        angle[0] = (angle[0] + 0.12) % (2 * math.pi)
        splash.update_idletasks()

    animation_id = [None]

    def animate():
        if not splash.winfo_exists():
            return

        draw_network_icon()
        splash.update()
        animation_id[0] = splash.after(30, animate)

    splash.deiconify()
    splash.update_idletasks()
    animation_id[0] = splash.after(30, animate)
    splash.update()
    splash._animated_id = animation_id
    return splash, label


# ============================================================
# Suppress Scapy pcap/libpcap warnings at startup
# ============================================================
logging.getLogger("scapy.runtime").setLevel(logging.ERROR)
os.environ["PYWARN"] = "ignore"

# Disable Scapy's warning printouts on stdout/stderr before import
sys.stderr_bak = sys.stderr
sys.stderr = open(os.devnull, 'w')

# ============================================================
# Auto-Install Missing Dependencies
# ============================================================

REQUIRED_PACKAGES = {
    "zeroconf": "zeroconf",
    "scapy": "scapy",
    "mac_vendor_lookup": "mac-vendor-lookup",
    "openpyxl": "openpyxl"
}

def install_missing_packages(splash_window=None, status_label=None):
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            __import__(module_name)
        except ImportError:
            missing.append(pip_name)

    if missing:
        if splash_window and status_label:
            status_label.config(text="Installing required libraries. Please wait...")
            splash_window.update()

        print(f"Missing required packages: {', '.join(missing)}")
        print("Automatically installing via pip...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
            print("Dependencies successfully installed!\n")
        except Exception as e:
            print(f"Error installing packages automatically: {e}")
            sys.exit(1)

    if splash_window and status_label:
        status_label.config(text="Loading network scanner components...")
        splash_window.update()


def close_startup_splash(splash_window):
    if splash_window is None:
        return

    try:
        animation_state = getattr(splash_window, "_animated_id", None)
        if animation_state is not None:
            animation_id = animation_state[0]
            if animation_id is not None:
                splash_window.after_cancel(animation_id)
    except Exception:
        pass

    try:
        if splash_window.winfo_exists():
            splash_window.update_idletasks()
            splash_window.destroy()
    except Exception:
        pass


import ipaddress
import socket
import concurrent.futures
import re
import time
import ctypes
import threading
import csv
from collections import Counter

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

try:
    from zeroconf import ServiceBrowser, Zeroconf
    import scapy.all as scapy
    from mac_vendor_lookup import MacLookup
finally:
    # Restore standard stderr after Scapy import finishes
    sys.stderr = sys.stderr_bak

def initialize_mac_lookup_engine():
    """Initialize live MAC lookup engine after splash displays so startup remains visible."""
    global mac_lookup_engine
    try:
        mac_lookup_engine = MacLookup()
        try:
            mac_lookup_engine.update_vendors()
        except Exception:
            pass
    except Exception:
        mac_lookup_engine = None


mac_lookup_engine = None


# ============================================================
# Persistent MAC Mapping Subsystem
# ============================================================

MAPPING_FILE = "mDNS-NetworkScanner-Mapping.json"

def load_mac_mappings():
    """Loads saved MAC address and note mappings from local JSON storage."""
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_mac_mapping(mac_address, ip_address, note_text):
    """Saves or updates a MAC/IP note mapping in the persistent JSON store."""
    mappings = load_mac_mappings()
    key = mac_address.upper() if mac_address and mac_address != "N/A" else f"IP:{ip_address}"
    
    mappings[key] = {
        "mac": mac_address,
        "last_ip": ip_address,
        "note": note_text,
        "updated": time.strftime("%Y-%m-%d %H:%M:%S")
    }
    
    try:
        with open(MAPPING_FILE, "w", encoding="utf-8") as f:
            json.dump(mappings, f, indent=4)
    except Exception as e:
        print(f"Error saving MAC mapping reference file: {e}")

def lookup_saved_note(mac_address, ip_address):
    """Retrieves a previously stored note for a MAC address or IP."""
    mappings = load_mac_mappings()
    if mac_address and mac_address != "N/A" and mac_address.upper() in mappings:
        return mappings[mac_address.upper()].get("note", "")
    
    ip_key = f"IP:{ip_address}"
    if ip_key in mappings:
        return mappings[ip_key].get("note", "")
        
    return ""


# ============================================================
# Global Settings & Tuning Defaults
# ============================================================

PING_WORKERS = 100
ARP_WORKERS = 50
HOSTNAME_WORKERS = 30
PORT_SCAN_WORKERS = 30

MDNS_SERVICE_DISCOVERY_TIME = 8
MDNS_DEVICE_DISCOVERY_TIME = 8

COMMON_PORTS = [21, 22, 23, 80, 135, 139, 443, 445, 3389, 8080]


# ============================================================
# Layer 3 ICMP Ping Engine
# ============================================================

def ping_host(ip_str, cancel_event=None):
    """Executes a fast ICMP echo request using native system ping."""
    if cancel_event and cancel_event.is_set():
        return False
    try:
        result = subprocess.run(
            ["ping", "-n", "1", "-w", "350", ip_str],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        return result.returncode == 0
    except Exception:
        return False


# ============================================================
# Native Win32 ARP Engine
# ============================================================

def send_arp_probe(ip_str, cancel_event=None):
    """
    Sends a direct Layer 2 ARP request via Windows iphlpapi.dll.
    Does NOT require Npcap/libpcap drivers to work reliably on Windows.
    """
    if cancel_event and cancel_event.is_set():
        return None
    try:
        inet_addr = socket.inet_aton(ip_str)
        ip_int = ctypes.c_ulong(int.from_bytes(inet_addr, byteorder='little'))
        
        mac_addr = (ctypes.c_byte * 6)()
        mac_len = ctypes.c_ulong(6)
        
        result = ctypes.windll.iphlpapi.SendARP(ip_int, 0, ctypes.byref(mac_addr), ctypes.byref(mac_len))
        
        if result == 0:
            bytes_mac = bytes(mac_addr)
            return ":".join(f"{b:02X}" for b in bytes_mac)
    except Exception:
        pass

    try:
        arp_request = scapy.ARP(pdst=ip_str)
        broadcast = scapy.Ether(dst="ff:ff:ff:ff:ff:ff")
        answered_list = scapy.srp(broadcast / arp_request, timeout=0.5, verbose=False)[0]
        if answered_list:
            return answered_list[0][1].hwsrc.upper()
    except Exception:
        pass

    return None


# ============================================================
# Network Detection & Range Parsing Engine
# ============================================================

def get_network_adapters():
    """Reads Windows ipconfig and returns active local IPv4 networks."""
    result = subprocess.run(
        ["ipconfig"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace"
    )

    networks = []
    current_adapter = None
    current_ip = None
    current_mask = None

    for raw_line in result.stdout.splitlines():
        line = raw_line.strip()

        adapter_match = re.match(r"^(.*adapter\s+)(.+):$", line, re.IGNORECASE)
        if adapter_match:
            current_adapter = adapter_match.group(2).strip()
            current_ip = None
            current_mask = None
            continue

        if "IPv4 Address" in line:
            match = re.search(r"(\d+\.\d+\.\d+\.\d+)", line)
            if match:
                current_ip = match.group(1)

        elif "Subnet Mask" in line:
            match = re.search(r"(\d+\.\d+\.\d+\.\d+)", line)
            if match:
                current_mask = match.group(1)

        if current_adapter and current_ip and current_mask:
            try:
                interface = ipaddress.ip_interface(f"{current_ip}/{current_mask}")
                network = interface.network

                if interface.ip.is_loopback:
                    current_ip = None
                    current_mask = None
                    continue

                existing = None
                for item in networks:
                    if item["network"] == network:
                        existing = item
                        break

                if existing:
                    if current_ip not in existing["ips"]:
                        existing["ips"].append(current_ip)
                    if current_adapter not in existing["adapters"]:
                        existing["adapters"].append(current_adapter)
                else:
                    networks.append({
                        "network": network,
                        "adapters": [current_adapter],
                        "ips": [current_ip]
                    })

                current_ip = None
                current_mask = None

            except ValueError:
                pass

    return networks


def parse_custom_range(user_input):
    """Parses custom CIDR notation, hyphenated ranges, or single IP strings."""
    user_input = user_input.strip()

    if "/" in user_input:
        net = ipaddress.ip_network(user_input, strict=False)
        return list(net.hosts()) if net.prefixlen < 32 else [net.network_address]

    if "-" in user_input:
        parts = user_input.split("-")
        start_str = parts[0].strip()
        end_str = parts[1].strip()

        start_ip = ipaddress.ip_address(start_str)

        if "." in end_str:
            end_ip = ipaddress.ip_address(end_str)
        else:
            base_octets = start_str.split(".")[:-1]
            end_ip = ipaddress.ip_address(".".join(base_octets + [end_str]))

        start_int = int(start_ip)
        end_int = int(end_ip)

        if start_int > end_int:
            start_int, end_int = end_int, start_int

        return [ipaddress.ip_address(ip) for ip in range(start_int, end_int + 1)]

    return [ipaddress.ip_address(user_input)]


import scapy.all as scapy
import time

def prewarm_arp_cache(target_ip):
    """
    Sends an explicit ARP request to pre-populate the local ARP cache 
    and ensure immediate MAC and OUI vendor resolution.
    """
    try:
        # Send a rapid, lightweight ARP request (timeout 0.5s)
        arp_req = scapy.Ether(dst="ff:ff:ff:ff:ff:ff") / scapy.ARP(pdst=str(target_ip))
        scapy.srp(arp_req, timeout=0.5, verbose=False)
    except Exception:
        pass

# ============================================================
# Hostname, NetBIOS, Live Vendor & Smart Fallback Engine
# ============================================================

def get_scapy_netbios(ip):
    """Fallback NetBIOS name query using Scapy packet crafting."""
    try:
        pkt = scapy.IP(dst=str(ip)) / scapy.UDP(sport=137, dport=137) / scapy.NBNSQueryRequest(SUFFIX="status", QUESTION_NAME="*")
        ans, _ = scapy.sr(pkt, timeout=1, verbose=False)
        for _, r in ans:
            if r.haslayer(scapy.NBNSQueryResponse):
                names = r[scapy.NBNSQueryResponse].ADDR_ENTRY
                for n in names:
                    name = n.RR_NAME.decode('ascii', errors='ignore').strip()
                    if name and not name.startswith("WORKGROUP"):
                        return name
    except Exception:
        pass
    return None


def get_netbios_name(ip):
    """Attempts NetBIOS hostname lookup using nbtstat with Scapy fallback."""
    try:
        result = subprocess.run(
            ["nbtstat", "-A", str(ip)],
            capture_output=True,
            text=True,
            errors="ignore",
            timeout=1.2
        )

        for line in result.stdout.splitlines():
            match = re.search(
                r"^\s*([A-Za-z0-9_\-\.]+)\s+<00>\s+UNIQUE",
                line,
                re.IGNORECASE
            )

            if match:
                name = match.group(1).strip()
                if name and not name.upper().startswith("WORKGROUP"):
                    return name

    except (subprocess.TimeoutExpired, subprocess.SubprocessError, OSError):
        pass

    # Restored Fallback
    return get_scapy_netbios(ip)


def resolve_host_info(ip, cancel_event=None):
    """Resolves DNS and NetBIOS hostname information."""
    if cancel_event and cancel_event.is_set():
        return ""
    try:
        hostname = socket.gethostbyaddr(str(ip))[0]
        if hostname and hostname != str(ip):
            return hostname
    except (socket.herror, socket.gaierror):
        pass

    netbios_name = get_netbios_name(ip)
    if netbios_name:
        return f"{netbios_name} (NetBIOS)"

    return ""


def get_vendor_by_mac(mac):
    """Performs live OUI Vendor Lookup."""
    if not mac or mac in ["N/A", "Gateway (Proxy)"] or not mac_lookup_engine:
        return ""
    try:
        vendor = mac_lookup_engine.lookup(mac)
        return vendor.strip()
    except Exception:
        pass
    return ""


def infer_smart_hostname(mac, open_ports_str):
    """Infers device class when hostnames cannot be resolved."""
    is_randomized_mac = False
    if mac and len(mac) >= 2:
        second_char = mac[1].upper()
        if second_char in ['2', '6', 'A', 'E']:
            is_randomized_mac = True

    ports = [p.strip() for p in open_ports_str.split(",") if p.strip().isdigit()]
    
    if "80" in ports or "443" in ports or "8080" in ports:
        return "[Web/Network Device]"
    if "135" in ports or "139" in ports or "445" in ports or "3389" in ports:
        return "[Windows Workstation/Server]"
    if "22" in ports:
        return "[Linux/SSH Host]"
    if "21" in ports or "23" in ports:
        return "[Legacy Device]"
    if is_randomized_mac:
        return "[Mobile/Private MAC Device]"

    return "[Active Host]"


# ============================================================
# ARP Table & Gateway Subsystem
# ============================================================

def get_arp_table():
    """Reads the system ARP cache."""
    result = subprocess.run(["arp", "-a"], capture_output=True, text=True)
    arp_devices = {}
    for line in result.stdout.splitlines():
        match = re.search(r"^\s*(\d+\.\d+\.\d+\.\d+)\s+([0-9a-fA-F-]{17})\s+dynamic", line)
        if match:
            ip = match.group(1)
            mac = match.group(2).replace("-", ":").upper()
            arp_devices[ip] = mac
    return arp_devices


def get_default_gateway_ip():
    """Returns the active default gateway IP for this host."""
    try:
        result = subprocess.run(
            ["route", "print", "0.0.0.0"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace"
        )
        for line in result.stdout.splitlines():
            if re.search(r"^\s*0\.0\.0\.0\s+\S+\s+(\d+\.\d+\.\d+\.\d+)", line, re.IGNORECASE):
                match = re.search(r"^\s*0\.0\.0\.0\s+\S+\s+(\d+\.\d+\.\d+\.\d+)", line, re.IGNORECASE)
                if match:
                    return match.group(1)
    except Exception:
        pass
    return None


def get_default_gateway_mac():
    """Identifies the default gateway's MAC address."""
    try:
        gw_ip = get_default_gateway_ip()
        if not gw_ip:
            return None
        arp_table = get_arp_table()
        return arp_table.get(gw_ip)
    except Exception:
        pass
    return None


# ============================================================
# mDNS Discovery Subsystem
# ============================================================

def mdns_discover(hosts_set, cancel_event=None, status_callback=None):
    """Executes active mDNS service discovery across target subnet."""
    if cancel_event and cancel_event.is_set():
        return {}

    if status_callback:
        status_callback("mDNS Discovery: Querying network service types...")

    service_types = set()

    def discover_service_types():
        discovered = set()
        zeroconf = Zeroconf()

        def service_type_handler(zeroconf=None, service_type=None, name=None, state_change=None, **kwargs):
            if name is None or state_change is None:
                return
            if getattr(state_change, "name", "") != "Added":
                return
            if name.endswith("._tcp.local.") or name.endswith("._udp.local."):
                discovered.add(name)

        browser = ServiceBrowser(zeroconf, "_services._dns-sd._udp.local.", handlers=[service_type_handler])
        
        for _ in range(MDNS_SERVICE_DISCOVERY_TIME * 10):
            if cancel_event and cancel_event.is_set():
                break
            time.sleep(0.1)

        browser.cancel()
        zeroconf.close()
        return discovered

    service_types.update(discover_service_types())

    if not service_types or (cancel_event and cancel_event.is_set()):
        return {}

    if status_callback:
        status_callback(f"mDNS Discovery: Querying {len(service_types)} service profiles...")

    mdns_devices = {}
    browsers = []
    zeroconf = Zeroconf()

    def service_handler(zeroconf=None, service_type=None, name=None, state_change=None, **kwargs):
        if cancel_event and cancel_event.is_set():
            return
        if name is None or service_type is None or state_change is None:
            return
        if getattr(state_change, "name", "") != "Added":
            return

        try:
            info = zeroconf.get_service_info(service_type, name, timeout=2000)
            if not info or not info.server:
                return

            hostname = info.server
            addresses = info.parsed_addresses()

            for address in addresses:
                try:
                    parsed_ip = ipaddress.ip_address(address)
                    if parsed_ip.version != 4 or address not in hosts_set:
                        continue
                except ValueError:
                    continue

                if address not in mdns_devices:
                    mdns_devices[address] = {"hostname": hostname, "services": set()}
                elif not mdns_devices[address]["hostname"]:
                    mdns_devices[address]["hostname"] = hostname

                mdns_devices[address]["services"].add(service_type)
        except Exception:
            pass

    for service_type in sorted(service_types):
        if cancel_event and cancel_event.is_set():
            break
        try:
            browser = ServiceBrowser(zeroconf, service_type, handlers=[service_handler])
            browsers.append(browser)
        except Exception:
            pass

    for _ in range(MDNS_DEVICE_DISCOVERY_TIME * 10):
        if cancel_event and cancel_event.is_set():
            break
        time.sleep(0.1)

    for browser in browsers:
        browser.cancel()
    zeroconf.close()

    return mdns_devices


# ============================================================
# TCP Port Audit Engine
# ============================================================

def scan_single_port(ip, port, cancel_event=None):
    """Connects to a single TCP port with tight timeout constraints."""
    if cancel_event and cancel_event.is_set():
        return None
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(0.25)
            if s.connect_ex((str(ip), port)) == 0:
                return port
    except Exception:
        pass
    return None


def scan_open_ports(ip, cancel_event=None):
    """Scans common TCP ports concurrently."""
    if cancel_event and cancel_event.is_set():
        return "None"
    open_ports = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(COMMON_PORTS)) as executor:
        futures = [executor.submit(scan_single_port, ip, port, cancel_event) for port in COMMON_PORTS]
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            if res:
                open_ports.append(res)
    open_ports.sort()
    return ", ".join(str(p) for p in open_ports) if open_ports else "None"


# ============================================================
# Complete Multi-Layer Scanner Driver
# ============================================================

def scan_network_gui(hosts, discovery_mode="fast", status_callback=None, cancel_event=None):
    hosts = list(hosts)
    hosts_set = set(str(ip) for ip in hosts)
    total = len(hosts)

    # 1. Multi-Threaded ICMP Ping Sweep
    if status_callback:
        status_callback(f"Scanning [Phase 1/5]: Multithreaded ICMP Ping Sweep ({total} targets)...")

    ping_hits = set()
    completed_pings = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=PING_WORKERS) as executor:
        futures = {executor.submit(ping_host, str(ip), cancel_event): str(ip) for ip in hosts}
        for future in concurrent.futures.as_completed(futures):
            if cancel_event and cancel_event.is_set():
                break
            ip = futures[future]
            completed_pings += 1
            if future.result():
                ping_hits.add(ip)
            if status_callback and completed_pings % 15 == 0:
                status_callback(f"Scanning [Phase 1/5]: ICMP Ping {completed_pings}/{total} ({completed_pings/total*100:.0f}%)")

    if cancel_event and cancel_event.is_set():
        return {}

    # 2. Layer-2 ARP Request Probing via Win32 API
    if status_callback:
        status_callback(f"Scanning [Phase 2/5]: Win32 Native ARP Probing...")

    direct_arp_hits = {}
    completed_arps = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=ARP_WORKERS) as executor:
        futures = {executor.submit(send_arp_probe, str(ip), cancel_event): str(ip) for ip in hosts}
        for future in concurrent.futures.as_completed(futures):
            if cancel_event and cancel_event.is_set():
                break
            ip = futures[future]
            completed_arps += 1
            try:
                mac = future.result()
                if mac:
                    direct_arp_hits[ip] = mac
            except Exception:
                pass

    if cancel_event and cancel_event.is_set():
        return {}

    system_arp = get_arp_table()
    combined_arp = {**system_arp, **direct_arp_hits}
    subnet_arp = {ip: mac for ip, mac in combined_arp.items() if ip in hosts_set}

    # RESTORED: Shared MAC / Proxy Analysis
    gateway_mac = get_default_gateway_mac()
    mac_counts = Counter(subnet_arp.values())
    proxy_macs = set()
    
    if gateway_mac:
        proxy_macs.add(gateway_mac)
        
    for mac, count in mac_counts.items():
        if count >= 3:
            proxy_macs.add(mac)

    # 3. Optional Active mDNS Sweep
    mdns_devices = {}
    if discovery_mode == "thorough" and not (cancel_event and cancel_event.is_set()):
        if status_callback:
            status_callback("Scanning [Phase 3/5]: Running mDNS Discovery...")
        mdns_devices = mdns_discover(hosts_set, cancel_event=cancel_event, status_callback=status_callback)

    if cancel_event and cancel_event.is_set():
        return {}

    raw_candidates = sorted(
        list(set(subnet_arp.keys()) | set(mdns_devices.keys()) | set(direct_arp_hits.keys())),
        key=ipaddress.ip_address
    )

    # 4. Hostname & NetBIOS Resolution
    if status_callback:
        status_callback(f"Scanning [Phase 4/5]: Resolving hostnames for {len(raw_candidates)} candidate(s)...")

    resolved_hostnames = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=HOSTNAME_WORKERS) as executor:
        futures = {executor.submit(resolve_host_info, ip, cancel_event): ip for ip in raw_candidates}
        for future in concurrent.futures.as_completed(futures):
            if cancel_event and cancel_event.is_set():
                break
            ip = futures[future]
            try:
                name = future.result()
                if name:
                    resolved_hostnames[ip] = name
            except Exception:
                pass

    if cancel_event and cancel_event.is_set():
        return {}

    # RESTORED: Ghost Filtering BEFORE Port Scanning
    active_target_ips = []
    for ip in raw_candidates:
        mac = subnet_arp.get(ip, "")
        is_proxy = mac in proxy_macs
        has_mdns = ip in mdns_devices
        has_hostname = ip in resolved_hostnames
        has_direct_arp = ip in direct_arp_hits

        # Proxy IPs require independent proof (mDNS, DNS/NetBIOS, or Direct ARP hit)
        if is_proxy:
            if has_mdns or has_hostname or has_direct_arp:
                active_target_ips.append(ip)
        else:
            active_target_ips.append(ip)

    # 5. TCP Port Scanning (Only Executed Against Validated Targets)
    if status_callback:
        status_callback(f"Scanning [Phase 5/5]: Auditing TCP ports for {len(active_target_ips)} real host(s)...")

    scanned_ports = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=PORT_SCAN_WORKERS) as executor:
        futures = {executor.submit(scan_open_ports, ip, cancel_event): ip for ip in active_target_ips}
        for future in concurrent.futures.as_completed(futures):
            if cancel_event and cancel_event.is_set():
                break
            ip = futures[future]
            try:
                scanned_ports[ip] = future.result()
            except Exception:
                scanned_ports[ip] = "None"

    if cancel_event and cancel_event.is_set():
        return {}

    # Consolidate Discovery Data
    devices = {}
    for ip in active_target_ips:
        mac = subnet_arp.get(ip, "")
        dns_netbios_name = resolved_hostnames.get(ip, "")
        mdns_name = mdns_devices.get(ip, {}).get("hostname", "")
        open_ports = scanned_ports.get(ip, "None")

        hostname = mdns_name or dns_netbios_name
        is_proxy_mac = mac in proxy_macs

        real_mac = mac if mac else ""
        vendor = get_vendor_by_mac(real_mac)

        if not hostname and vendor:
            hostname = f"[{vendor}]"

        if not hostname:
            hostname = infer_smart_hostname(real_mac, open_ports)

        methods = []
        if ip in ping_hits:
            methods.append("ICMP Ping")
        if ip in direct_arp_hits:
            methods.append("ARP Direct")
        elif mac:
            methods.append("ARP Cache")
        if ip in mdns_devices:
            methods.append("mDNS")
        if dns_netbios_name:
            methods.append("DNS/NetBIOS")
        if open_ports != "None":
            methods.append("TCP Port")

        saved_note = lookup_saved_note(real_mac, ip)

        # Determine vendor string format
        if is_proxy_mac:
            vendor_label = f"{vendor if vendor else 'Unknown Vendor'} [Gateway/Proxy]"
        else:
            vendor_label = vendor if vendor else "Unknown Vendor"

        devices[ip] = {
            "mac": mac if mac else "N/A",
            "vendor": vendor_label,
            "hostname": hostname,
            "method": " + ".join(methods) if methods else "Active Probe",
            "ports": open_ports,
            "notes": saved_note,
            "discovered_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

    return devices


# ============================================================
# Graphical User Interface (Tkinter)
# ============================================================
class TreeviewTooltip:
    """Displays a hover tooltip over Treeview cells."""
    def __init__(self, tree):
        self.tree = tree
        self.tip_window = None

    def show_tip(self, text, x, y):
        if self.tip_window or not text:
            return
        self.tip_window = tw = tk.Toplevel(self.tree)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x+15}+{y+15}")
        label = tk.Label(
            tw, text=text, justify=tk.LEFT,
            background="#FFFFE1", relief=tk.SOLID, borderwidth=1,
            font=("Segoe UI", 8, "normal"), padx=4, pady=2
        )
        label.pack()

    def hide_tip(self):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


class NetworkScannerGUI:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1360x760")
        self.root.minsize(1040, 580)
        self.apply_window_branding()

        self.adapters = get_network_adapters()
        self.scan_results = {}
        self.current_display_results = {}
        self._scan_previous_results = {}
        self.cancel_event = threading.Event()
        self.is_scan_running = False
        self._has_completed_scan = False

        self.setup_styles()
        self.create_widgets()

        # Bind double click for editing notes
        self.tree.bind("<Double-1>", self.on_double_click)

        # --- ADD THESE THREE LINES ---
        self.tooltip = TreeviewTooltip(self.tree)
        self.tree.bind("<Motion>", self.on_tree_hover)
        self.tree.bind("<Button-3>", self.show_context_menu)
        # -----------------------------

    def on_tree_hover(self, event):
        """Shows a hover tooltip on treeview rows."""
        item_id = self.tree.identify_row(event.y)
        if item_id:
            x = self.root.winfo_pointerx()
            y = self.root.winfo_pointery()
            self.tooltip.show_tip("💡 Right-click row/cell to copy to clipboard", x, y)
        else:
            self.tooltip.hide_tip()

    def show_context_menu(self, event):
        """Creates a right-click context menu to copy selected row/cell data."""
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        
        if not item_id:
            return

        self.tree.selection_set(item_id)
        values = self.tree.item(item_id, "values")
        
        # Determine column index clicked (e.g., '#1' -> 0)
        col_index = int(column_id.replace("#", "")) - 1 if column_id else 0
        cell_value = values[col_index] if 0 <= col_index < len(values) else ""

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(
            label=f"Copy Cell ('{cell_value[:20]}...')", 
            command=lambda: self.copy_to_clipboard(cell_value)
        )
        menu.add_command(
            label="Copy Entire Row (CSV Format)", 
            command=lambda: self.copy_to_clipboard(", ".join(f'"{v}"' for v in values))
        )
        menu.post(event.x_root, event.y_root)

    def copy_to_clipboard(self, text):
        """Copies text to the system clipboard."""
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        self.set_status(f"Copied to clipboard: '{text}'", status_type="ready")

    def apply_window_branding(self):
        """Applies the same title and icon used by the splash window."""
        self.root.title("mDNS Network Scanner")
        self.root.configure(bg="#F8FAFC")
        brand_icon = get_branding_icon()
        if brand_icon:
            try:
                self.root.iconphoto(False, brand_icon)
            except Exception:
                pass

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        self.style.configure("Treeview", font=("Segoe UI", 9), rowheight=24)
        self.style.configure("ScanPrimary.TButton", background="#166534", foreground="white", font=("Segoe UI", 9, "bold"))
        self.style.map("ScanPrimary.TButton", background=[("active", "#14532d"), ("disabled", "#4b5563")], foreground=[("active", "white"), ("disabled", "#e5e7eb")])
        self.style.configure("RescanOrange.TButton", background="#C2410C", foreground="white", font=("Segoe UI", 9, "bold"))
        self.style.map("RescanOrange.TButton", background=[("active", "#9A4B0A"), ("disabled", "#9ca3af")], foreground=[("active", "white"), ("disabled", "#f3f4f6")])
        self.style.configure("StopRed.TButton", background="#7F1D1D", foreground="white", font=("Segoe UI", 9, "bold"))
        self.style.map("StopRed.TButton", background=[("active", "#450A0A"), ("disabled", "#9ca3af")], foreground=[("active", "white"), ("disabled", "#f3f4f6")])
        self.style.configure("ExportBlue.TButton", background="#240EE9", foreground="white", font=("Segoe UI", 9, "bold"))
        self.style.map("ExportBlue.TButton", background=[("active", "#0284C7"), ("disabled", "#94a3b8")], foreground=[("active", "white"), ("disabled", "#f8fafc")])

        self.brand_icon = get_branding_icon()

    def create_widgets(self):
        branding_bar = tk.Frame(self.root, bg="#E2E8F0", height=34)
        branding_bar.pack(fill=tk.X, padx=0, pady=(0, 4))
        branding_bar.pack_propagate(False)

        icon_label = tk.Label(branding_bar, bg="#E2E8F0", bd=0, padx=6, pady=2)
        icon_label.pack(side=tk.LEFT, padx=(10, 0))
        if self.brand_icon:
            icon_label.configure(image=self.brand_icon)
            icon_label.image = self.brand_icon

        title_label = tk.Label(
            branding_bar,
            text="mDNS Network Scanner",
            bg="#E2E8F0",
            fg="#0F172A",
            font=("Segoe UI", 13, "bold"),
            anchor=tk.W,
        )
        title_label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 1. Top Configuration Frame
        ctrl_frame = ttk.LabelFrame(self.root, text=" Network Selection & Configuration ", padding=8)
        ctrl_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(ctrl_frame, text="Interface / Subnet:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=(2, 1))
        self.adapter_var = tk.StringVar()
        self.adapter_combo = ttk.Combobox(ctrl_frame, textvariable=self.adapter_var, width=50, state="readonly")

        combo_values = []
        for net in self.adapters:
            net_str = str(net['network'])
            ip_str = ", ".join(net['ips'])
            adapter_str = ", ".join(net['adapters'])
            combo_values.append(f"{net_str} ({ip_str}) - {adapter_str}")

        combo_values.append("Custom IP / Range / CIDR...")
        self.adapter_combo['values'] = combo_values
        if combo_values:
            self.adapter_combo.current(0)
        self.adapter_combo.grid(row=0, column=1, sticky=tk.W, padx=5, pady=(2, 1))
        self.adapter_combo.bind("<<ComboboxSelected>>", self.on_adapter_select)

        ttk.Label(ctrl_frame, text="Custom Range:").grid(row=0, column=2, sticky=tk.W, padx=(12, 5), pady=(2, 1))
        self.custom_entry = ttk.Entry(ctrl_frame, width=28, state="disabled")
        self.custom_entry.grid(row=0, column=3, sticky=tk.W, padx=5, pady=(2, 1))

        ttk.Label(ctrl_frame, text="Discovery Mode:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=(2, 1))
        self.mode_var = tk.StringVar(value="thorough")
        ttk.Radiobutton(ctrl_frame, text="Fast Scan (Ping + ARP + NetBIOS + Ports)", variable=self.mode_var, value="fast").grid(row=1, column=1, sticky=tk.W, padx=5, pady=(2, 1))
        ttk.Radiobutton(ctrl_frame, text="Thorough Scan (Fast + Full mDNS)", variable=self.mode_var, value="thorough").grid(row=1, column=2, columnspan=2, sticky=tk.W, padx=(12, 5), pady=(2, 1))

        # Action Buttons Box
        btn_box = ttk.Frame(ctrl_frame)
        btn_box.grid(row=0, column=4, rowspan=2, padx=10, pady=2, sticky="NSEW")

        self.scan_btn = ttk.Button(btn_box, text="Start Scan", style="ScanPrimary.TButton", command=lambda: self.start_scan_thread(preserve_existing=False), width=12)
        self.scan_btn.pack(side=tk.TOP, fill=tk.X, pady=2)

        self.rescan_btn = ttk.Button(btn_box, text="Rescan", style="RescanOrange.TButton", command=lambda: self.start_scan_thread(preserve_existing=True), width=12, state="disabled")
        self.rescan_btn.pack(side=tk.TOP, fill=tk.X, pady=2)

        self.stop_btn = ttk.Button(btn_box, text="Stop Scan", style="StopRed.TButton", command=self.stop_scan, state="disabled", width=12)
        self.stop_btn.pack(side=tk.TOP, fill=tk.X, pady=2)

        # 2. Live Search & Filter Bar (unified 50/50 split)
        filter_frame = ttk.Frame(self.root, padding=(8, 3))
        filter_frame.pack(fill=tk.X, padx=10, pady=2)

        # Left half: Search & Filter controls
        search_panel = ttk.Frame(filter_frame)
        search_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        ttk.Label(search_panel, text="Filter Target:").pack(side=tk.LEFT, padx=(0, 3))
        self.filter_field_var = tk.StringVar(value="All Fields")
        self.filter_combo = ttk.Combobox(
            search_panel,
            textvariable=self.filter_field_var,
            values=["All Fields", "IP Address", "MAC Address", "Vendor", "Hostname", "Open Ports", "Notes"],
            state="readonly",
            width=14
        )
        self.filter_combo.pack(side=tk.LEFT, padx=3)
        self.filter_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_filter())

        ttk.Label(search_panel, text="Query:").pack(side=tk.LEFT, padx=(10, 3))
        self.filter_query_var = tk.StringVar()
        self.filter_entry = ttk.Entry(search_panel, textvariable=self.filter_query_var, width=28)
        self.filter_entry.pack(side=tk.LEFT, padx=3)
        self.filter_query_var.trace_add("write", lambda *args: self.apply_filter())

        clear_filter_btn = ttk.Button(search_panel, text="Clear", command=self.clear_filter, width=8)
        clear_filter_btn.pack(side=tk.LEFT, padx=5)

        # Right half: Export controls
        export_panel = ttk.Frame(filter_frame)
        export_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(8, 0))

        self.export_btn = ttk.Button(export_panel, text="Export Excel", style="ExportBlue.TButton", command=self.export_to_excel)
        self.export_btn.pack(fill=tk.BOTH, expand=True)

        # 3. Results Table Frame
        tree_frame = ttk.Frame(self.root, padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("ip", "mac", "vendor", "hostname", "method", "ports", "discovered", "notes")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")

        self.tree.heading("ip", text="IP Address")
        self.tree.heading("mac", text="MAC Address")
        self.tree.heading("vendor", text="Vendor / Manufacturer")
        self.tree.heading("hostname", text="Hostname / Resolved Device")
        self.tree.heading("method", text="Discovery Method")
        self.tree.heading("ports", text="Open Ports")
        self.tree.heading("discovered", text="Discovered")
        self.tree.heading("notes", text="Notes (Double-Click to Edit / Auto-Saved)")

        self.tree.column("ip", width=110, anchor=tk.W)
        self.tree.column("mac", width=130, anchor=tk.W)
        self.tree.column("vendor", width=160, anchor=tk.W)
        self.tree.column("hostname", width=200, anchor=tk.W)
        self.tree.column("method", width=160, anchor=tk.W)
        self.tree.column("ports", width=110, anchor=tk.W)
        self.tree.column("discovered", width=150, anchor=tk.W)
        self.tree.column("notes", width=240, anchor=tk.W)

        self.tree.tag_configure("existing", background="white")
        self.tree.tag_configure("new", background="#D9F7E8")
        self.tree.tag_configure("missing", background="#E5E7EB")

        self.sort_state = {}
        for col in columns:
            self.sort_state[col] = "none"
            self.tree.heading(col, command=lambda c=col: self.sort_treeview(c))

        self.tree.bind("<Double-1>", self.on_double_click)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscroll=scrollbar_y.set, xscroll=scrollbar_x.set)

        self.tree.grid(row=0, column=0, sticky="NSEW")
        scrollbar_y.grid(row=0, column=1, sticky="NS")
        scrollbar_x.grid(row=1, column=0, sticky="EW")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 4. Color-Coded Status Bar
        self.status_bar = tk.Label(
            self.root,
            text=" Status: Ready | Reference file: mDNS-NetworkScanner-Mapping.json",
            bg="#2E7D32",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            anchor=tk.W,
            padx=10,
            pady=6
        )
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def set_status(self, text, status_type="running"):
        """
        Updates status bar background and text.
        status_type: "running", "stopping", "stopped", "ready"
        """
        color_map = {
            "running": "#E65100",
            "stopping": "#D97706",
            "stopped": "#B91C1C",
            "ready": "#2E7D32"
        }
        bg_color = color_map.get(status_type, "#2E7D32")
        
        def update():
            self.status_bar.config(text=f" Status: {text}", bg=bg_color)
        self.root.after(0, update)

    def on_adapter_select(self, event):
        if self.adapter_combo.get() == "Custom IP / Range / CIDR...":
            self.custom_entry.config(state="normal")
            self.custom_entry.focus()
        else:
            self.custom_entry.config(state="disabled")

    def update_action_button_states(self):
        self.scan_btn.config(state="disabled" if self.is_scan_running else "normal")
        self.stop_btn.config(state="normal" if self.is_scan_running else "disabled")
        self.rescan_btn.config(state="normal" if (self._has_completed_scan and not self.is_scan_running) else "disabled")

    def stop_scan(self):
        """Triggers cancellation event and updates status indicator."""
        if self.is_scan_running:
            self.cancel_event.set()
            self.update_action_button_states()
            self.set_status("Cancellation requested... Waiting for active thread pools to finalize...", status_type="stopping")

    def update_sort_indicator(self, column, direction):
        """Sets a visible ascending/descending arrow on the active sorted column."""
        column_labels = {
            "ip": "IP Address",
            "mac": "MAC Address",
            "vendor": "Vendor / Manufacturer",
            "hostname": "Hostname / Resolved Device",
            "method": "Discovery Method",
            "ports": "Open Ports",
            "discovered": "Discovered",
            "notes": "Notes (Double-Click to Edit / Auto-Saved)"
        }

        for key in self.sort_state:
            self.sort_state[key] = "none"

        self.sort_state[column] = direction

        for col_key, label in column_labels.items():
            current_state = self.sort_state.get(col_key, "none")
            arrow = " ▲" if current_state == "asc" else " ▼" if current_state == "desc" else ""
            self.tree.heading(col_key, text=label + arrow)

    def sort_treeview(self, column):
        """Sort treeview rows by selected column value."""
        items = [(self.tree.set(child, column), child) for child in self.tree.get_children("")]

        def sort_key(entry):
            value, child = entry
            text = str(value or "").strip()
            if column == "ip":
                try:
                    return (0, int(ipaddress.ip_address(text)))
                except ValueError:
                    return (1, text.lower())
            if column == "discovered":
                try:
                    return (0, datetime.datetime.strptime(text, "%Y-%m-%d %H:%M:%S"))
                except ValueError:
                    return (1, text.lower())
            if column == "ports":
                try:
                    if text == "None":
                        return (0, -1)
                    first = next(int(part) for part in text.split(",") if part.strip().isdigit())
                    return (0, first)
                except StopIteration:
                    return (1, text.lower())
            return (0, text.lower())

        current_direction = self.sort_state.get(column, "none")
        if current_direction == "asc":
            direction = "desc"
            items.sort(key=sort_key, reverse=True)
        else:
            direction = "asc"
            items.sort(key=sort_key)

        for index, (_, child) in enumerate(items):
            self.tree.move(child, "", index)

        self.update_sort_indicator(column, direction)

    def start_scan_thread(self, preserve_existing=False):
        selected = self.adapter_combo.get()
        if not selected:
            messagebox.showwarning("Selection Warning", "Please select a network interface or custom range.")
            return

        try:
            if selected == "Custom IP / Range / CIDR...":
                raw_input = self.custom_entry.get().strip()
                if not raw_input:
                    messagebox.showwarning("Input Error", "Please specify a target IP, range, or CIDR block.")
                    return
                hosts = parse_custom_range(raw_input)
            else:
                idx = self.adapter_combo.current()
                hosts = list(self.adapters[idx]['network'].hosts())
        except Exception as e:
            messagebox.showerror("Parsing Error", f"Failed to parse target network range: {e}")
            return

        if not hosts:
            messagebox.showwarning("Target Error", "No valid target host IP addresses generated.")
            return

        if not preserve_existing:
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.current_display_results = {}
        else:
            self._scan_previous_results = dict(self.scan_results)

        self.cancel_event.clear()
        self.is_scan_running = True
        self.update_action_button_states()
        self.set_status("Initializing native multi-stage network scanner...", status_type="running")

        mode = self.mode_var.get()

        threading.Thread(
            target=self.run_scan,
            args=(hosts, mode),
            daemon=True
        ).start()

    def run_scan(self, hosts, mode):
        """Executes the network scan in a background thread."""
        gateway_ip = get_default_gateway_ip()

        # Pre-warm the ARP cache for the actual gateway before parsing devices.
        if gateway_ip:
            prewarm_arp_cache(gateway_ip)

        try:
            results = scan_network_gui(
                hosts,
                discovery_mode=mode,
                status_callback=lambda msg: self.set_status(msg, status_type="running") if not self.cancel_event.is_set() else None,
                cancel_event=self.cancel_event
            )

            previous_results = self._scan_previous_results if hasattr(self, "_scan_previous_results") else {}
            if self.cancel_event.is_set():
                self._has_completed_scan = False
                self.scan_results = {}
                self.root.after(0, self.populate_results, {}, previous_results)
                self.set_status("Scan stopped mid-operation by user. Results cleared.", status_type="stopped")
            else:
                self._has_completed_scan = True
                self.scan_results = results
                self.root.after(0, self.populate_results, results, previous_results)
        except Exception as e:
            self._has_completed_scan = False
            self.root.after(0, lambda: messagebox.showerror("Scan Failure", str(e)))
            self.set_status(f"Scan aborted due to error: {e}", status_type="stopped")
        finally:
            self.is_scan_running = False
            self.root.after(0, self.update_action_button_states)

    def populate_results(self, results, previous_results=None):
        if previous_results is None:
            previous_results = {}

        merged_results = dict(previous_results)
        merged_results.update(results)
        self.current_display_results = merged_results

        for item in self.tree.get_children():
            self.tree.delete(item)

        for ip in sorted(merged_results.keys(), key=ipaddress.ip_address):
            info = merged_results[ip]
            if ip in results and ip in previous_results:
                tag = "existing"
            elif ip in results and ip not in previous_results:
                tag = "new"
            elif ip in previous_results and ip not in results:
                tag = "missing"
            else:
                tag = "existing"

            row_data = {
                "IP Address": ip,
                "MAC Address": info.get('mac', 'N/A'),
                "Vendor": info.get('vendor', 'Unknown Vendor'),
                "Hostname": info.get('hostname', ''),
                "Open Ports": info.get('ports', 'None'),
                "Discovered": info.get('discovered_at', "Unknown"),
                "Notes": info.get('notes', '')
            }

            self.tree.insert(
                "",
                tk.END,
                values=(
                    ip,
                    row_data["MAC Address"],
                    row_data["Vendor"],
                    row_data["Hostname"],
                    info.get('method', 'Active Probe'),
                    row_data["Open Ports"],
                    row_data["Discovered"],
                    row_data["Notes"]
                ),
                tags=(tag,)
            )

        self.apply_filter()
        if not results and not self.cancel_event.is_set():
            self.set_status("Scan completed. No active hosts discovered on subnet.", status_type="ready")
        elif results:
            self.set_status(f"Scan completed successfully. Discovered {len(results)} active host(s).", status_type="ready")

    # ============================================================
    # Live Search & Filter Logic
    # ============================================================

    def clear_filter(self):
        self.filter_query_var.set("")
        self.filter_field_var.set("All Fields")
        self.apply_filter()

    def apply_filter(self):
        query = self.filter_query_var.get().strip().lower()
        field_target = self.filter_field_var.get()

        for item in self.tree.get_children():
            self.tree.delete(item)

        source_results = self.current_display_results if self.current_display_results else self.scan_results
        if not source_results:
            return

        for ip in sorted(source_results.keys(), key=ipaddress.ip_address):
            info = source_results[ip]
            row_data = {
                "IP Address": ip,
                "MAC Address": info.get('mac', 'N/A'),
                "Vendor": info.get('vendor', 'Unknown Vendor'),
                "Hostname": info.get('hostname', ''),
                "Open Ports": info.get('ports', 'None'),
                "Discovered": info.get('discovered_at', 'Unknown'),
                "Notes": info.get('notes', '')
            }

            match = False
            if not query:
                match = True
            elif field_target == "All Fields":
                match = any(query in str(val).lower() for val in [ip, info.get('mac', 'N/A'), info.get('vendor', 'Unknown Vendor'), info.get('hostname', ''), info.get('method', 'Active Probe'), info.get('ports', 'None'), info.get('notes', ''), info.get('discovered_at', 'Unknown')])
            elif field_target in row_data:
                match = query in str(row_data[field_target]).lower()

            if match:
                tag = "existing"
                if ip in self.scan_results and ip not in self._scan_previous_results:
                    tag = "new"
                elif ip not in self.scan_results and ip in self._scan_previous_results:
                    tag = "missing"

                self.tree.insert(
                    "",
                    tk.END,
                    values=(
                        ip,
                        row_data["MAC Address"],
                        row_data["Vendor"],
                        row_data["Hostname"],
                        info.get('method', 'Active Probe'),
                        row_data["Open Ports"],
                        row_data["Discovered"],
                        row_data["Notes"]
                    ),
                    tags=(tag,)
                )

    def on_double_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        current_values = list(self.tree.item(item_id, "values"))
        ip_addr = current_values[0]
        mac_addr = current_values[1]
        current_note = current_values[6] if len(current_values) > 6 else ""

        new_note = simpledialog.askstring(
            "Edit Device Note",
            f"Enter custom note for {ip_addr} ({mac_addr}):",
            initialvalue=current_note,
            parent=self.root
        )

        if new_note is not None:
            cleaned_note = new_note.strip()
            current_values[6] = cleaned_note
            self.tree.item(item_id, values=current_values)
            
            if ip_addr in self.scan_results:
                self.scan_results[ip_addr]["notes"] = cleaned_note
            
            save_mac_mapping(mac_addr, ip_addr, cleaned_note)
            self.set_status(f"Saved note for {ip_addr} ({mac_addr}) to mapping reference file.", status_type="ready")

    def export_to_excel(self):
        if not self.scan_results and not self.tree.get_children():
            messagebox.showwarning("Export Warning", "No scan data available to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Spreadsheet", "*.xlsx"), ("CSV File", "*.csv")],
            title="Export Scan Results"
        )

        if not file_path:
            return

        headers = ["IP Address", "MAC Address", "Vendor / Manufacturer", "Hostname / Resolved Device", "Discovery Method", "Open Ports", "Discovered", "Notes"]
        rows = []

        for item_id in self.tree.get_children():
            rows.append(self.tree.item(item_id, "values"))

        try:
            if file_path.endswith(".xlsx"):
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Network Audit"

                ws.append(headers)
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                for row_data in rows:
                    ws.append(row_data)

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                wb.save(file_path)

            else:
                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)

            messagebox.showinfo("Export Successful", f"Scan results exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not write export file: {e}")


# ============================================================
# Main Program Launch Point
# ============================================================

def main():
    splash = show_startup_splash("Please wait, loading libraries...")
    splash_window, splash_label = splash

    def bootstrap_in_background():
        try:
            install_missing_packages(splash_window, splash_label)
            initialize_mac_lookup_engine()
        finally:
            if splash_window.winfo_exists():
                splash_window.after(0, lambda: close_startup_splash(splash_window))

    startup_thread = threading.Thread(target=bootstrap_in_background, daemon=True)
    startup_thread.start()

    splash_window.mainloop()

    root = tk.Tk()
    root.withdraw()
    root.update_idletasks()

    app = NetworkScannerGUI(root)
    root.update_idletasks()
    root.deiconify()
    root.mainloop()


if __name__ == "__main__":
    main()